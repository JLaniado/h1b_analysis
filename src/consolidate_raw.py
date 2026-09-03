"""
Consolidate the raw DOL exports (mixed .csv and .xlsx, split across fiscal
years/quarters) into two single, clean, structured CSVs — one for LCA
(H-1B/E-3/H-1B1), one for PERM — that everything downstream reads from.

Why this exists: DOL publishes these files inconsistently across years —
sometimes one cumulative year-to-date CSV (FY2026 Q3), sometimes four
separate per-quarter XLSX exports (FY2025), each internally padded with
blank template rows out to a fixed row count, with a stray column-naming
difference between years (`H-1B_DEPENDENT` vs `H_1B_DEPENDENT`). One
quarter (FY2025 Q2) also exports nearly every text cell as a literal
Excel "keep as text" wrapper (`="541511"` instead of `541511`), which
silently breaks anything that parses those columns (e.g. SOC_CODE's
major-group prefix) if left in place. None of that should leak into
analysis code — this module absorbs it once, here.

Output (gzip-compressed — pandas reads .csv.gz natively, no decompression
step needed, and it's ~4x smaller, which matters for distributing this data
as a GitHub Release asset; see src/fetch_master_data.py):
  data/interim/lca_master.csv.gz   (all populated LCA/H-1B records, every source)
  data/interim/perm_master.csv.gz  (all populated PERM records, every source)

Both gain three columns not present in any raw source:
  FISCAL_YEAR, FISCAL_QUARTER — derived from DECISION_DATE (falls back to
    RECEIVED_DATE when a case has no decision yet), using the federal fiscal
    calendar (FY starts Oct 1; Q1=Oct-Dec, Q2=Jan-Mar, Q3=Apr-Jun, Q4=Jul-Sep).
    Derived rather than trusted from the filename, since raw files aren't
    reliably disjoint by decision quarter (see module docstring history in
    git log / notebook 00 for the investigation).
  SOURCE_FILE — original filename, for traceability back to a raw export.

These are large files (multiple GB) and are gitignored (data/interim/) —
re-run this script against data/raw/ to reproduce them.
"""

import csv
import datetime
import gzip
import re
from pathlib import Path

import openpyxl

# One source file (LCA_Disclosure_Data_FY2025_Q2.xlsx) exports nearly every
# text cell wrapped as an Excel "keep as text" formula literal, e.g.
# `="541511"` instead of `541511`. Strip that wrapper wherever it shows up —
# cheap to check, and protects any future source that carries the same
# artifact. Only matches a *whole-cell* wrap, not a legitimate value that
# happens to contain "=" internally.
_EXCEL_TEXT_WRAP_RE = re.compile(r'^="(.*)"$')

RAW_DIR = Path(__file__).resolve().parent.parent / "data" / "raw"
INTERIM_DIR = Path(__file__).resolve().parent.parent / "data" / "interim"

LCA_SOURCES = [
    "LCA_Disclosure_Data_FY2025_Q1.xlsx",
    "LCA_Disclosure_Data_FY2025_Q2.xlsx",
    "LCA_Disclosure_Data_FY2025_Q3.xlsx",
    "LCA_Disclosure_Data_FY2025_Q4.xlsx",
    "LCA_Disclosure_Data_FY2026_Q3.csv",
]

PERM_SOURCES = [
    "PERM_Disclosure_Data_FY2025_Q4.xlsx",
    "PERM_Disclosure_Data_FY2026_Q3.csv",
]

# Column name has changed across vintages; normalize to one name.
COLUMN_RENAMES = {
    "H-1B_DEPENDENT": "H_1B_DEPENDENT",
}

EXTRA_COLUMNS = ["FISCAL_YEAR", "FISCAL_QUARTER", "SOURCE_FILE"]


def _fiscal_year_quarter(d: datetime.date) -> tuple[int, int]:
    """Federal fiscal year/quarter for a given date (FY starts Oct 1)."""
    if d.month >= 10:
        fy = d.year + 1
    else:
        fy = d.year
    quarter = {10: 1, 11: 1, 12: 1, 1: 2, 2: 2, 3: 2, 4: 3, 5: 3, 6: 3, 7: 4, 8: 4, 9: 4}[d.month]
    return fy, quarter


def _parse_date(value):
    """Normalize a raw date cell (datetime, or MM/DD/YY string) to a date, or None."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    for fmt in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.datetime.strptime(str(value), fmt).date()
        except ValueError:
            continue
    return None


def _iso(d):
    return d.isoformat() if d else ""


def _unwrap_excel_text(value):
    if isinstance(value, str):
        m = _EXCEL_TEXT_WRAP_RE.match(value)
        if m:
            return m.group(1)
    return value


def _rows_from_xlsx(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [COLUMN_RENAMES.get(h, h) for h in next(rows)]
    for row in rows:
        yield {k: _unwrap_excel_text(v) for k, v in zip(header, row)}
    wb.close()


def _rows_from_csv(path: Path):
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        reader.fieldnames = [COLUMN_RENAMES.get(h, h) for h in reader.fieldnames]
        for row in reader:
            yield {k: _unwrap_excel_text(v) for k, v in row.items()}


def _winning_source_by_case(sources: list[str], case_number_col: str) -> dict:
    """First pass: for any CASE_NUMBER appearing in more than one source file,
    the *last* source in the list wins (sources are listed oldest-to-newest,
    so this prefers the most recently published snapshot of a case's status
    if the raw files turn out to be overlapping/cumulative rather than
    disjoint by quarter)."""
    winner: dict[str, str] = {}
    for name in sources:
        path = RAW_DIR / name
        rows_iter = _rows_from_xlsx(path) if path.suffix == ".xlsx" else _rows_from_csv(path)
        for row in rows_iter:
            case_no = row.get(case_number_col)
            if case_no:
                winner[case_no] = name
    return winner


def _consolidate(sources: list[str], out_path: Path, case_number_col: str = "CASE_NUMBER"):
    all_fieldnames: list[str] = []
    for name in sources:
        path = RAW_DIR / name
        rows_iter = _rows_from_xlsx(path) if path.suffix == ".xlsx" else _rows_from_csv(path)
        first_row = next(rows_iter, None)
        if first_row is None:
            continue
        for k in first_row.keys():
            if k not in all_fieldnames:
                all_fieldnames.append(k)

    fieldnames = all_fieldnames + EXTRA_COLUMNS
    INTERIM_DIR.mkdir(parents=True, exist_ok=True)

    print("  scanning for duplicate case numbers across sources...")
    winning_source = _winning_source_by_case(sources, case_number_col)

    total_written = 0
    total_skipped_stale_duplicate = 0
    total_blank = 0

    with gzip.open(out_path, "wt", newline="", encoding="utf-8") as out_f:
        writer = csv.DictWriter(out_f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()

        for name in sources:
            path = RAW_DIR / name
            print(f"  reading {name} ...")
            rows_iter = _rows_from_xlsx(path) if path.suffix == ".xlsx" else _rows_from_csv(path)
            written_from_source = 0

            for row in rows_iter:
                case_no = row.get(case_number_col)
                if not case_no:
                    total_blank += 1
                    continue
                if winning_source[case_no] != name:
                    total_skipped_stale_duplicate += 1
                    continue

                decision = _parse_date(row.get("DECISION_DATE"))
                received = _parse_date(row.get("RECEIVED_DATE"))
                fy, fq = _fiscal_year_quarter(decision or received)

                out_row = dict(row)
                for date_col in ("RECEIVED_DATE", "DECISION_DATE", "ORIGINAL_CERT_DATE"):
                    if date_col in out_row:
                        out_row[date_col] = _iso(_parse_date(out_row.get(date_col)))
                out_row["FISCAL_YEAR"] = fy
                out_row["FISCAL_QUARTER"] = fq
                out_row["SOURCE_FILE"] = name

                writer.writerow(out_row)
                written_from_source += 1
                total_written += 1

            print(f"    -> {written_from_source:,} records")

    print(f"Wrote {total_written:,} records to {out_path}")
    print(f"Dropped {total_blank:,} blank template rows, skipped {total_skipped_stale_duplicate:,} stale duplicate case numbers")


def build_lca_master():
    print("Building LCA master...")
    _consolidate(LCA_SOURCES, INTERIM_DIR / "lca_master.csv.gz")


def build_perm_master():
    print("Building PERM master...")
    _consolidate(PERM_SOURCES, INTERIM_DIR / "perm_master.csv.gz")


if __name__ == "__main__":
    build_lca_master()
    build_perm_master()
